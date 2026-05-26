from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "lib" / "extracted-workbook-data.json"
PDF_TEXT_OUT = ROOT / "src" / "lib" / "source-plan-decision-text.txt"


def norm(value) -> str:
    return unicodedata.normalize("NFC", str(value)).strip() if value is not None else ""


def slug(text: str) -> str:
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-zA-Z0-9]+", "-", text.lower()).strip("-")


def username_from_name(full_name: str) -> str:
    cleaned = full_name
    for prefix in ("Ông ", "Bà ", "Ồng "):
        cleaned = cleaned.replace(prefix, "")
    cleaned = unicodedata.normalize("NFD", cleaned)
    cleaned = "".join(ch for ch in cleaned if unicodedata.category(ch) != "Mn")
    parts = [p.lower() for p in re.findall(r"[A-Za-z0-9]+", cleaned)]
    if not parts:
        return "user"
    return f"{parts[-1]}{''.join(p[0] for p in parts[:-1])}"


def role_from_source(role: str) -> str:
    if "Thư ký" in role:
        return "Thư ký đoàn"
    if "Phó" in role:
        return "Phó trưởng đoàn"
    if "Trưởng đoàn" in role or "Trưởng" in role:
        return "Trưởng đoàn"
    if "Chỉ đạo" in role:
        return "Ban Giám đốc"
    return "Thành viên đoàn"


def extract_pdf_text() -> str:
    pdf = next(ROOT.glob("*.pdf"))
    reader = PdfReader(str(pdf))
    text = "\n".join(
        f"\n--- PAGE {index + 1} ---\n{page.extract_text() or ''}"
        for index, page in enumerate(reader.pages)
    )
    PDF_TEXT_OUT.write_text(text, encoding="utf-8")
    return text


def extract_workbooks():
    workbooks = []
    departments = []
    criteria_sets = {}
    criteria_items = []
    form_templates = []
    form_criteria_items = []
    team_members_by_name = {}
    seen_departments = set()
    seen_items = set()

    for path in sorted(ROOT.glob("*.xlsx")):
        team = "Đoàn 01" if "ĐOÀN 1" in path.name else "Đoàn 02"
        form_type = "LS-CLS" if "LS-CLS" in path.name else "Hành chính"
        wb = load_workbook(path, data_only=True, read_only=False)
        forms = []

        for ws in wb.worksheets:
            if ws.title == "PHAN_CONG_02_DOAN":
                header_row = 2 if (ws.cell(1, 1).value or "").startswith("PHÂN") else 1
                for row in range(header_row + 1, (ws.max_row or header_row) + 1):
                    audit_team = norm(ws.cell(row, 1).value)
                    full_name = norm(ws.cell(row, 2).value)
                    if not audit_team or not full_name:
                        continue
                    team_members_by_name[full_name] = {
                        "id": f"user-{slug(full_name)}",
                        "username": username_from_name(full_name),
                        "fullName": full_name,
                        "titleUnit": norm(ws.cell(row, 3).value),
                        "role": role_from_source(norm(ws.cell(row, 4).value)),
                        "auditTeam": audit_team,
                        "sourceRole": norm(ws.cell(row, 4).value),
                        "assignedGroups": [item.strip() for item in norm(ws.cell(row, 5).value).split(",") if item.strip()],
                        "email": "",
                        "phone": "",
                        "active": True,
                        "defaultPassword": "123456",
                        "dataSource": "Quyết định 271/QĐ-BV và sheet PHAN_CONG_02_DOAN"
                    }
                continue

            is_admin = ws.title.startswith("P_")
            is_lscls = ws.title.startswith("LS_") or ws.title.startswith("CLS_")
            if not (is_admin or is_lscls):
                continue

            if is_admin:
                department_name = norm(ws.cell(4, 2).value)
                block = "Hành chính"
                header = 6
                mapping = {"stt": 1, "ma": 2, "group": 3, "content": 4, "evidence": 5, "max": 6, "team1": 13, "team2": 14}
                set_id = "criteria-hanh-chinh-v03"
                form_type_for_sheet = "Hành chính"
            else:
                department_name = norm(ws.cell(4, 8).value)
                block = "Cận lâm sàng" if ws.title.startswith("CLS_") else "Lâm sàng"
                header = 9
                mapping = {"stt": 1, "ma": 2, "group": 3, "content": 4, "max": 5, "evidence": 8, "team1": 9, "team2": 10}
                set_id = "criteria-ls-cls-v03"
                form_type_for_sheet = "LS-CLS"

            if not department_name:
                continue

            department_id = slug(department_name)
            if department_id not in seen_departments:
                departments.append({
                    "id": department_id,
                    "name": department_name,
                    "block": block,
                    "active": True,
                    "dataSource": "Kế hoạch 32/KH-BV và các sheet phiếu kiểm tra"
                })
                seen_departments.add(department_id)

            rows = []
            total_score = 0
            form_template_id = f"form-template-{slug(path.stem)}-{slug(ws.title)}"
            form_type_code = "HANH_CHINH" if is_admin else "LS_CLS"
            for row in range(header + 1, (ws.max_row or header) + 1):
                stt = ws.cell(row, mapping["stt"]).value
                if not isinstance(stt, (int, float)):
                    continue
                content = norm(ws.cell(row, mapping["content"]).value)
                item = {
                    "id": f"{set_id}-{int(stt):02d}",
                    "criteriaSetId": set_id,
                    "order": int(stt),
                    "groupCode": norm(ws.cell(row, mapping["ma"]).value),
                    "groupName": norm(ws.cell(row, mapping["group"]).value),
                    "content": content,
                    "maxScore": float(ws.cell(row, mapping["max"]).value or 0),
                    "scoringType": "nhập điểm",
                    "isHighRisk": any(
                        keyword in content.lower()
                        for keyword in ["cấp cứu", "thuốc", "nhiễm khuẩn", "an toàn", "hồ sơ", "bảo mật", "an ninh", "tài sản"]
                    ),
                    "evidenceHint": norm(ws.cell(row, mapping["evidence"]).value),
                    "assignedGroupTeam1": norm(ws.cell(row, mapping["team1"]).value),
                    "assignedGroupTeam2": norm(ws.cell(row, mapping["team2"]).value),
                    "dataSource": "File Excel bảng kiểm V03_1805"
                }
                rows.append(item)
                total_score += item["maxScore"]
                if item["id"] not in seen_items:
                    criteria_items.append(item)
                    seen_items.add(item["id"])
                form_criteria_items.append({
                    "id": f"{form_template_id}-row-{int(stt):02d}",
                    "formTemplateId": form_template_id,
                    "criteriaId": item["id"],
                    "sourceFile": path.name,
                    "sourceSheet": ws.title,
                    "sourceRow": row,
                    "formType": form_type_code,
                    "departmentCode": ws.title,
                    "departmentName": department_name,
                    "inspectionTeam": team,
                    "version": "V03-1805",
                    "order": int(stt),
                    "groupCode": item["groupCode"],
                    "groupName": item["groupName"],
                    "content": item["content"],
                    "evidenceRequired": item["evidenceHint"],
                    "maxScore": item["maxScore"],
                    "team1Assignee": item["assignedGroupTeam1"],
                    "team2Assignee": item["assignedGroupTeam2"],
                    "isHighRiskRelated": item["isHighRisk"]
                })

            if set_id not in criteria_sets:
                criteria_sets[set_id] = {
                    "id": set_id,
                    "formType": form_type_for_sheet,
                    "applicableTeams": ["Đoàn 01", "Đoàn 02"],
                    "totalScore": total_score,
                    "criteriaCount": len(rows),
                    "version": "V03-1805",
                    "active": True,
                    "dataSource": "Kế hoạch 32/KH-BV và file Excel bảng kiểm"
                }

            forms.append({
                "sheet": ws.title,
                "department": department_name,
                "block": block,
                "criteriaSetId": set_id,
                "criteriaCount": len(rows),
                "totalScore": total_score
            })
            form_templates.append({
                "id": form_template_id,
                "name": f"Phiếu kiểm tra và chấm điểm - {department_name}",
                "sourceFile": path.name,
                "sourceSheet": ws.title,
                "sourceRow": 1,
                "formType": form_type_code,
                "departmentId": department_id,
                "departmentCode": ws.title,
                "departmentName": department_name,
                "block": block,
                "inspectionTeam": team,
                "version": "V03-1805",
                "totalScore": total_score,
                "criteriaCount": len(rows),
                "headerFields": [
                    {"key": "ten_phieu", "label": "Tên phiếu", "value": f"Phiếu kiểm tra và chấm điểm - {department_name}", "sourceCell": "A1"},
                    {"key": "don_vi_duoc_kiem_tra", "label": "Đơn vị được kiểm tra", "value": department_name, "sourceCell": "B4" if is_admin else "H4"},
                    {"key": "khoi", "label": "Khối/khoa/phòng", "value": block, "sourceCell": "derived"},
                    {"key": "thang_diem", "label": "Thang điểm", "value": str(int(total_score)), "sourceCell": "derived"},
                    {"key": "so_noi_dung", "label": "Số nội dung kiểm tra", "value": str(len(rows)), "sourceCell": "derived"},
                    {"key": "doan_kiem_tra", "label": "Đoàn kiểm tra", "value": team, "sourceCell": "derived"},
                    {"key": "ngay_kiem_tra", "label": "Ngày kiểm tra", "value": "Chọn theo phiên kiểm tra", "sourceCell": "runtime"},
                    {"key": "nguoi_tiep_doan", "label": "Người tiếp đoàn", "value": "Nhập khi tạo phiếu", "sourceCell": "runtime"},
                    {"key": "thoi_gian", "label": "Thời gian bắt đầu/kết thúc", "value": "Nhập khi tạo phiếu", "sourceCell": "runtime"},
                    {"key": "ket_luan_so_bo", "label": "Kết luận sơ bộ", "value": "Nhập sau kiểm tra", "sourceCell": "runtime"}
                ]
            })

        workbooks.append({
            "file": path.name,
            "team": team,
            "formType": form_type,
            "sheetCount": len(wb.worksheets),
            "formCount": len(forms),
            "forms": forms
        })

    users = list(team_members_by_name.values())
    users.insert(0, {
        "id": "user-admin-demo",
        "username": "admin",
        "fullName": "Quản trị hệ thống",
        "titleUnit": "Tài khoản mẫu giai đoạn prototype",
        "role": "Admin",
        "auditTeam": "Toàn viện",
        "sourceRole": "Admin mẫu",
        "assignedGroups": ["Tất cả"],
        "email": "",
        "phone": "",
        "active": True,
        "defaultPassword": "123456",
        "dataSource": "Dữ liệu mẫu - chờ cấu hình tài khoản thật"
    })
    users.insert(1, {
        "id": "user-khth-demo",
        "username": "khth",
        "fullName": "Tài khoản Phòng Kế hoạch tổng hợp",
        "titleUnit": "Phòng Kế hoạch tổng hợp",
        "role": "Phòng KHTH",
        "auditTeam": "Toàn viện",
        "sourceRole": "Đầu mối quản trị dữ liệu, tổng hợp báo cáo",
        "assignedGroups": ["Quản trị dữ liệu", "Tổng hợp báo cáo", "Chốt kỳ kiểm tra"],
        "email": "",
        "phone": "",
        "active": True,
        "defaultPassword": "123456",
        "dataSource": "Kế hoạch 32/KH-BV quy định Phòng KHTH là đầu mối; tài khoản là dữ liệu mẫu"
    })
    users.append({
        "id": "user-capa-demo",
        "username": "capa",
        "fullName": "Đầu mối CAPA khoa/phòng",
        "titleUnit": "Tài khoản mẫu cập nhật khắc phục",
        "role": "CAPA",
        "auditTeam": "",
        "sourceRole": "CAPA mẫu",
        "assignedGroups": ["Khắc phục"],
        "email": "",
        "phone": "",
        "active": True,
        "defaultPassword": "123456",
        "dataSource": "Dữ liệu mẫu - chờ cấu hình đầu mối từng khoa/phòng"
    })

    return workbooks, departments, list(criteria_sets.values()), criteria_items, users, form_templates, form_criteria_items


AUDIT_SCHEDULE = [
    ("2026-05-27", ["Hồi sức tích cực - Chống độc Nhi", "Cấp cứu - Chống độc Sản"], "Quản lý chất lượng", ["Sơ sinh", "Nội tổng hợp"], "Dược"),
    ("2026-06-03", ["Sanh", "Sản thường"], "Tổ chức cán bộ", ["Truyền nhiễm", "Ngoại nhi"], "Xét nghiệm"),
    ("2026-06-10", ["Phụ", "Sản bệnh"], "Hành chính quản trị", ["Liên chuyên khoa", "Hiếm muộn"], "Kiểm soát nhiễm khuẩn"),
    ("2026-06-17", ["Hậu phẫu", "Gây mê hồi sức"], "Điều dưỡng", ["Khám bệnh", "Cấp cứu Nhi"], "Chẩn đoán hình ảnh"),
    ("2026-06-24", ["Sơ sinh", "Nội tổng hợp"], "Tài chính kế toán", ["Hồi sức tích cực - Chống độc Nhi", "Cấp cứu - Chống độc Sản"], "Dinh dưỡng"),
    ("2026-07-01", ["Truyền nhiễm", "Ngoại nhi"], "Vật tư thiết bị y tế", ["Sanh", "Sản thường"], "Kế hoạch tổng hợp"),
    ("2026-07-08", ["Liên chuyên khoa", "Hiếm muộn"], "Công tác xã hội", ["Phụ", "Sản bệnh"], "Quản lý chất lượng"),
    ("2026-07-15", ["Khám bệnh", "Cấp cứu Nhi"], "Dược", ["Hậu phẫu", "Gây mê hồi sức"], "Tổ chức cán bộ"),
    ("2026-07-22", ["Hồi sức tích cực - Chống độc Nhi", "Cấp cứu - Chống độc Sản"], "Xét nghiệm", ["Sơ sinh", "Nội tổng hợp"], "Hành chính quản trị"),
    ("2026-07-29", ["Sanh", "Sản thường"], "Kiểm soát nhiễm khuẩn", ["Truyền nhiễm", "Ngoại nhi"], "Điều dưỡng"),
    ("2026-08-05", ["Phụ", "Sản bệnh"], "Chẩn đoán hình ảnh", ["Liên chuyên khoa", "Hiếm muộn"], "Tài chính kế toán"),
    ("2026-08-12", ["Hậu phẫu", "Gây mê hồi sức"], "Dinh dưỡng", ["Khám bệnh", "Cấp cứu Nhi"], "Vật tư thiết bị y tế"),
    ("2026-08-19", ["Sơ sinh", "Nội tổng hợp"], "Kế hoạch tổng hợp", ["Hồi sức tích cực - Chống độc Nhi", "Cấp cứu - Chống độc Sản"], "Công tác xã hội"),
    ("2026-08-26", ["Truyền nhiễm", "Ngoại nhi"], "Quản lý chất lượng", ["Sanh", "Sản thường"], "Dược"),
    ("2026-09-02", ["Liên chuyên khoa", "Hiếm muộn"], "Tổ chức cán bộ", ["Phụ", "Sản bệnh"], "Xét nghiệm"),
    ("2026-09-09", ["Khám bệnh", "Cấp cứu Nhi"], "Hành chính quản trị", ["Hậu phẫu", "Gây mê hồi sức"], "Kiểm soát nhiễm khuẩn"),
    ("2026-09-16", ["Hồi sức tích cực - Chống độc Nhi", "Cấp cứu - Chống độc Sản"], "Điều dưỡng", ["Sơ sinh", "Nội tổng hợp"], "Chẩn đoán hình ảnh"),
    ("2026-09-23", ["Sanh", "Sản thường"], "Tài chính kế toán", ["Truyền nhiễm", "Ngoại nhi"], "Dinh dưỡng"),
    ("2026-09-30", ["Phụ", "Sản bệnh"], "Vật tư thiết bị y tế", ["Liên chuyên khoa", "Hiếm muộn"], "Kế hoạch tổng hợp"),
    ("2026-10-07", ["Hậu phẫu", "Gây mê hồi sức"], "Công tác xã hội", ["Khám bệnh", "Cấp cứu Nhi"], "Quản lý chất lượng"),
    ("2026-10-14", ["Sơ sinh", "Nội tổng hợp"], "Dược", ["Hồi sức tích cực - Chống độc Nhi", "Cấp cứu - Chống độc Sản"], "Tổ chức cán bộ"),
    ("2026-10-21", ["Truyền nhiễm", "Ngoại nhi"], "Xét nghiệm", ["Sanh", "Sản thường"], "Hành chính quản trị"),
    ("2026-10-28", ["Liên chuyên khoa", "Hiếm muộn"], "Kiểm soát nhiễm khuẩn", ["Phụ", "Sản bệnh"], "Điều dưỡng"),
    ("2026-11-04", ["Khám bệnh", "Cấp cứu Nhi"], "Chẩn đoán hình ảnh", ["Hậu phẫu", "Gây mê hồi sức"], "Tài chính kế toán"),
    ("2026-11-11", ["Hồi sức tích cực - Chống độc Nhi", "Cấp cứu - Chống độc Sản"], "Dinh dưỡng", ["Sơ sinh", "Nội tổng hợp"], "Vật tư thiết bị y tế"),
    ("2026-11-18", ["Sanh", "Sản thường"], "Kế hoạch tổng hợp", ["Truyền nhiễm", "Ngoại nhi"], "Công tác xã hội"),
    ("2026-11-25", ["Phụ", "Sản bệnh"], "Quản lý chất lượng", ["Liên chuyên khoa", "Hiếm muộn"], "Dược"),
    ("2026-12-02", ["Hậu phẫu", "Gây mê hồi sức"], "Tổ chức cán bộ", ["Khám bệnh", "Cấp cứu Nhi"], "Xét nghiệm"),
    ("2026-12-09", ["Sơ sinh", "Nội tổng hợp"], "Hành chính quản trị", ["Hồi sức tích cực - Chống độc Nhi", "Cấp cứu - Chống độc Sản"], "Kiểm soát nhiễm khuẩn"),
    ("2026-12-16", ["Truyền nhiễm", "Ngoại nhi"], "Điều dưỡng", ["Sanh", "Sản thường"], "Chẩn đoán hình ảnh"),
    ("2026-12-23", ["Liên chuyên khoa", "Hiếm muộn"], "Tài chính kế toán", ["Phụ", "Sản bệnh"], "Dinh dưỡng"),
    ("2026-12-30", ["Khám bệnh", "Cấp cứu Nhi"], "Vật tư thiết bị y tế", ["Hậu phẫu", "Gây mê hồi sức"], "Kế hoạch tổng hợp"),
]


def main():
    extract_pdf_text()
    workbooks, departments, criteria_sets, criteria_items, users, form_templates, form_criteria_items = extract_workbooks()
    output = {
        "sourceFiles": [w["file"] for w in workbooks],
        "sourceDocuments": [
            {
                "name": "Kế hoạch số 32/KH-BV",
                "file": next(ROOT.glob("*.pdf")).name,
                "usedFor": ["Danh mục khoa/phòng", "lịch kiểm tra", "quy định chấm điểm", "quy trình báo cáo/CAPA"]
            },
            {
                "name": "Quyết định số 271/QĐ-BV",
                "file": next(ROOT.glob("*.pdf")).name,
                "usedFor": ["Thành viên đoàn", "vai trò", "nhiệm vụ chấm", "nguyên tắc hoạt động"]
            },
            {
                "name": "Các file Excel bảng kiểm V03_1805",
                "file": "*.xlsx",
                "usedFor": ["Bộ phiếu", "tiêu chí", "minh chứng", "người phụ trách từng tiêu chí"]
            }
        ],
        "dataProvenance": [
            {"area": "Khoa/phòng, khối", "source": "Kế hoạch 32/KH-BV + sheet phiếu Excel", "status": "Dữ liệu thật từ nguồn"},
            {"area": "Đoàn kiểm tra, thành viên, nhiệm vụ", "source": "Quyết định 271/QĐ-BV + PHAN_CONG_02_DOAN", "status": "Dữ liệu thật từ nguồn"},
            {"area": "Bảng kiểm, tiêu chí, minh chứng", "source": "4 file Excel bảng kiểm", "status": "Dữ liệu thật từ nguồn"},
            {"area": "Username/password", "source": "Prototype", "status": "Dữ liệu mẫu chờ cấu hình sau"},
            {"area": "Upload file minh chứng", "source": "Prototype", "status": "Chưa triển khai, hiện lưu text/ghi chú"},
            {"area": "Định dạng Excel chính thức", "source": "Prototype", "status": "Mẫu mặc định chờ chuẩn hóa biểu mẫu cuối"}
        ],
        "workbookSummary": workbooks,
        "departments": departments,
        "users": users,
        "criteriaSets": criteria_sets,
        "criteriaItems": criteria_items,
        "formTemplates": form_templates,
        "formCriteriaItems": form_criteria_items,
        "auditSchedule": [
            {
                "id": f"schedule-{index + 1:02d}",
                "auditDate": date,
                "team1ClinicalDepartments": team1_clinical,
                "team1SupportDepartment": team1_support,
                "team2ClinicalDepartments": team2_clinical,
                "team2SupportDepartment": team2_support,
                "note": "Chấm trên Smartphone; thư ký tổng hợp ngay sau kiểm tra",
                "dataSource": "Phụ lục lịch kiểm tra Kế hoạch 32/KH-BV"
            }
            for index, (date, team1_clinical, team1_support, team2_clinical, team2_support) in enumerate(AUDIT_SCHEDULE)
        ],
        "defaultRules": {
            "scoreScale": 100,
            "lsClsCriteriaCount": 30,
            "adminCriteriaCount": 20,
            "classifications": [
                {"name": "Tốt", "min": 90, "max": 100},
                {"name": "Đạt", "min": 80, "max": 89.99},
                {"name": "Cần cải tiến", "min": 65, "max": 79.99},
                {"name": "Không đạt", "min": 0, "max": 64.99}
            ],
            "forceFailWhenSevereRisk": True,
            "highRiskGroups": ["cấp cứu", "thuốc", "kiểm soát nhiễm khuẩn", "an toàn người bệnh", "an toàn thông tin", "an ninh trật tự", "hồ sơ bệnh án", "tài sản công"],
            "evidenceModePrototype": "text",
            "reportColumns": [
                "Ngày kiểm tra", "Tháng", "Năm", "Đơn vị", "Khối", "Đoàn kiểm tra", "Người chấm", "Nhóm tiêu chí", "Nội dung tiêu chí", "Điểm tối đa", "Điểm đạt", "Điểm trừ", "Tổng điểm", "Xếp loại", "Minh chứng", "Nội dung khắc phục", "Người chịu trách nhiệm", "Thời hạn hoàn thành", "Trạng thái CAPA", "Mức độ nguy cơ"
            ],
            "dataSource": "Mục VI, VII Kế hoạch 32/KH-BV"
        }
    }
    OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "departments": len(departments),
        "users": len(users),
        "criteriaSets": len(criteria_sets),
        "criteriaItems": len(criteria_items),
        "scheduleRows": len(AUDIT_SCHEDULE)
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
